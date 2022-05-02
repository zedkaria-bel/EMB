import pandas as pd
import numpy as np
import re
import datetime
import os, sys
from openpyxl import load_workbook
import dateparser
import psycopg2
import sqlalchemy
from sqlalchemy import create_engine
from sqlalchemy import func
from sqlalchemy.orm import sessionmaker

# BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# sys.path.append(BASE_DIR)
# os.environ['DJANGO_SETTINGS_MODULE'] = 'core.settings'
# os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

con = psycopg2.connect(database='EMB', user='postgres', password='zaki1690', host='localhost', port='5432')
con.autocommit = True
engine = create_engine(f'postgresql://postgres:zaki1690@localhost:5432/EMB')
Session = sessionmaker(bind = engine)
session = Session()
cursor = con.cursor()

pd.set_option('display.max_rows', None)
pd.set_option('use_inf_as_na', True)

full_labels = ['prep_line',
'pause_eat',
'chg_form',
'lvg',
'manque_prog',
'panne',
'reglages',
'autres',
'abs']

file_str = r"C:\Users\zaki1\Desktop\Controle de Gestion\Scripts\PRODUCTION V2\FLASH JOURNALIER D'IMPRESSION 07 04 2022.xlsx"

book = load_workbook(file_str)
xl = pd.ExcelFile(file_str, engine='openpyxl') # pylint: disable=abstract-class-instantiated
bg_frames = []
for sheetname in xl.sheet_names:
    if 'flash' in sheetname.lower() or 'journ' in sheetname.lower():
        bg_df = pd.read_excel(file_str, sheet_name=sheetname, header=1)
        ix = bg_df.index[bg_df['Unnamed: 0'].str.lower().str.contains(r'\d{2}/\d{2}/\d{4}', na = False)].tolist()[0]
        dte = bg_df['Unnamed: 0'][ix]
        # check the data type
        if isinstance(dte, datetime.datetime):
            date = dte.date()
        else:
            match = re.search(r'\d{2}/\d{2}/\d{4}', dte)
            date = datetime.datetime.strptime(match.group(), '%d/%m/%Y').date()
        dfs = bg_df.index[bg_df['Unnamed: 0'].str.lower().str.contains('vern|tand|offs', na = False)].tolist()
        for idx, val in enumerate(dfs):
            df = pd.DataFrame()
            try:
                df = bg_df.loc[val: dfs[idx+1]]
            except IndexError:
                df = bg_df.loc[val:]
            df = df.reset_index(drop=True)
            line = df.iloc[0]['Unnamed: 0']
            if 'vern' in line.lower():
                line = 'VERNISSEUSE'
            else:
                line = 'TANDEM'
            # **************************************** FLASH IMPRESSION ***********************************************************
            end_flash_df = df.index[df['Unnamed: 3'].str.lower().str.contains('tot', na = False)].tolist()[0]
            flash_df = df.loc[0:end_flash_df, :]
            flash_df = flash_df.copy()
            flash_df.dropna(axis=1, how='all', inplace = True)
            to_rmv = df.index[df['Unnamed: 0'].str.lower().str.contains('shi', na = False)].tolist()[0]
            flash_df = flash_df.iloc[to_rmv + 3:-1]
            flash_df['Unnamed: 0'].fillna(method='ffill', inplace = True)
            flash_df['Unnamed: 1'].fillna(method='ffill', inplace = True)
            flash_df['Unnamed: 13'].fillna(method='ffill', inplace = True)
            flash_df['Unnamed: 0'] = flash_df['Unnamed: 0'].str.strip().str[-1].astype(int)
            # flash_df = flash_df.iloc[: , 1:]
            flash_df.rename(columns = {
                flash_df.columns[0]: 'hours',
                flash_df.columns[1]: 'shift',
                flash_df.columns[2]: 'format_fer',
                flash_df.columns[3]: 'des',
                flash_df.columns[4]: 'nb_psg',
                flash_df.columns[5]: 'sf_brut',
                flash_df.columns[6]: 'sf_rebut',
                flash_df.columns[7]: 'sf_conf',
                flash_df.columns[8]: 'sf_taux_reb',
                flash_df.columns[9]: 'brut',
                flash_df.columns[10]: 'rebut',
                flash_df.columns[11]: 'conf',
                flash_df.columns[12]: 'taux_reb',
                flash_df.columns[13]: 'conduct',
            }, inplace = True)
            flash_df = flash_df.reset_index(drop=True)
            flash_df = flash_df.replace(['-'],0)
            flash_df.fillna(0, inplace = True)
            flash_df['sf_brut'] = flash_df['sf_rebut'] + flash_df['sf_conf']
            try:
                flash_df['sf_taux_reb'] = flash_df['sf_rebut'] / flash_df['sf_brut']
            except ZeroDivisionError():
                flash_df['sf_taux_reb'] = np.nan
            flash_df['brut'] = flash_df['rebut'] + flash_df['conf']
            try:
                flash_df['taux_reb'] = flash_df['rebut'] / flash_df['brut']
            except ZeroDivisionError():
                flash_df['taux_reb'] = np.nan
            flash_df['date'] = date
            flash_df['ligne'] = line
            print(flash_df)
            # **************************************** FLASH IMPRESSION ***********************************************************
            
            # **************************************** CAPACITE PROD ***********************************************************
            # start_df2 = df.index[df['Unnamed: 0'].str.lower().str.contains('arr', na = False)].tolist()[0]
            # try:
            #     df_capacite = df.loc[start_df2:dfs[idx+1], :]
            # except:
            #     df_capacite = df.loc[start_df2:, :]
            # df_capacite = df_capacite.reset_index(drop=True)
            # df_capacite.dropna(axis=1, how='all', inplace = True)
            # df_arrets = df_capacite.iloc[:, :3]
            # df_cap_prod = df_capacite.iloc[:, 3:]
            
            # # DF ARRETS

            # begin_df = (df_arrets['Unnamed: 0'].str.lower().str.contains('prog', na = False)).idxmax()
            # end_df = df_arrets['Unnamed: 2'].where(df_arrets['Unnamed: 2'].str.lower().str.contains('tot', na = False)).last_valid_index()
            # df_arrets = df_arrets.loc[begin_df:end_df-1]
            # df_arrets = df_arrets.loc[:, 'Unnamed: 2':]
            # df_arrets.rename(columns = {
            #     df_arrets.columns[0]: 'label_arret',
            #     df_arrets.columns[1]: 'temps arrets (mn)',
            # }, inplace = True)
            # df_arrets = df_arrets[df_arrets['label_arret'].notna()]
            # # df_arrets['line'] = line
            # # df_arrets['date'] = date

            # df_arrets.loc[df_arrets['label_arret'].str.lower().str.contains('parat'), 'label_arret'] = 'prep_line'
            # df_arrets.loc[df_arrets['label_arret'].str.lower().str.contains('paus'), 'label_arret'] = 'pause_eat'
            # df_arrets.loc[df_arrets['label_arret'].str.lower().str.contains('forma'), 'label_arret'] = 'chg_form'
            # df_arrets.loc[df_arrets['label_arret'].str.lower().str.contains('lavag'), 'label_arret'] = 'lvg'
            # df_arrets.loc[df_arrets['label_arret'].str.lower().str.contains('prog'), 'label_arret'] = 'manque_prog'
            # df_arrets.loc[df_arrets['label_arret'].str.lower().str.contains('pan'), 'label_arret'] = 'panne'
            # df_arrets.loc[df_arrets['label_arret'].str.lower().str.contains('reg|rég'), 'label_arret'] = 'reglages'
            # df_arrets.loc[df_arrets['label_arret'].str.lower().str.contains('autr'), 'label_arret'] = 'autres'
            # df_arrets.loc[df_arrets['label_arret'].str.lower().str.contains('abs'), 'label_arret'] = 'abs'

            # diff = list(set(full_labels) - set(df_arrets['label_arret'].unique().tolist()))
            # if len(diff) > 0:
            #     for label in diff:
            #         df_arrets.loc[len(df_arrets.index)+1] = [label, 0]

            # df_arrets.fillna(0, inplace = True)
            # df_arrets = df_arrets.T
            # df_arrets.columns = df_arrets.iloc[0]
            # df_arrets = df_arrets.iloc[1:]
            # # print(df_arrets)

            # # DF CAP PROD
        
            # df_cap_prod.rename(columns = {
            #     df_cap_prod.columns[0]: 'Unnamed: 0',
            #     df_cap_prod.columns[1]: 'Unnamed: 1',
            # }, inplace = True)

            # cph_brut = df_cap_prod.index[df_cap_prod['Unnamed: 0'].str.lower().str.contains('=', na = False)].tolist()[0]
            # match = re.search(r'([0-9]+)\s*f', df_cap_prod.at[cph_brut, 'Unnamed: 0'])
            # cph = int(match.group(1))
            # df_cap_prod = df_cap_prod.iloc[1:]
            # df_cap_prod = df_cap_prod.dropna(axis=0, how='all')
            # df_cap_prod.rename(columns = {
            #     df_cap_prod.columns[0]: 'key',
            #     df_cap_prod.columns[1]: 'val',
            # }, inplace = True)
            # df_cap_prod = df_cap_prod[df_cap_prod['val'].notna()]
            # df_cap_prod.loc[df_cap_prod['key'].str.lower().str.contains('arr'), 'key'] = 'arrets'
            # df_cap_prod.loc[df_cap_prod['key'].str.lower().str.contains('bru'), 'key'] = 'prod_brute'
            # df_cap_prod.loc[df_cap_prod['key'].str.lower().str.contains('shif'), 'key'] = 'shift'
            # df_cap_prod.loc[df_cap_prod['key'].str.lower().str.contains('util'), 'key'] = 'taux_util'
            # df_cap_prod.loc[df_cap_prod['key'].str.lower().str.contains('cap'), 'key'] = 'capacite_prod'
            # df_cap_prod = df_cap_prod.T
            # df_cap_prod.columns = df_cap_prod.iloc[0]
            # df_cap_prod = df_cap_prod.iloc[1:]
            # s = df_cap_prod.columns.to_series()
            # s.iloc[-1] = 'taux_prod'
            # df_cap_prod.columns = s

            # # COMBINE DF_ARRET AND DF_CAP_PROD
            # df_arrets.reset_index(drop=True, inplace=True)
            # df_cap_prod.reset_index(drop=True, inplace=True)
            # df_capacite = pd.concat([df_cap_prod, df_arrets], axis=1)
            # df_capacite['date'] = date
            # df_capacite['ligne'] = line
            # df_capacite['cph'] = cph

            # **************************************** CAPACITE PROD ***********************************************************

            # print(flash_df)
            # print(df_capacite)

            # tab_name = 'Production_Capacite_Imp'

            # cursor.execute('SELECT MAX("id") FROM public."' + tab_name + '"')
            # max_id = cursor.fetchone()[0]
            # if not max_id:
            #     max_id = 0
            # df_capacite.insert(0, 'id', range(int(max_id) + 1, 1 + int(max_id) + len(df_capacite)))
            # df_capacite.to_sql(tab_name, engine, if_exists='append', index=False)

            tab_name = 'Flash_Impression'

            max_id = None
            try:
                cursor.execute('SELECT MAX("id") FROM public."' + tab_name + '"')
                max_id = cursor.fetchone()[0]
            except:
                pass
            if not max_id:
                max_id = 0
            flash_df.insert(0, 'id', range(int(max_id) + 1, 1 + int(max_id) + len(flash_df)))
            flash_df.to_sql(tab_name, engine, if_exists='append', index=False)



# ************************** OLD CAPACITE PRODUCTION VERNISSEUSE TANDEM ****************************


# book = load_workbook(file_str)
# xl = pd.ExcelFile(file_str, engine='openpyxl') # pylint: disable=abstract-class-instantiated
# bg_frames = []
# for sheetname in xl.sheet_names:
#     # TANDEM or VERNIS. Sheets
#     bg_df = pd.read_excel(file_str, sheet_name=sheetname, header=1)
#     if 'vern' in sheetname.strip().lower() or 'tand' in sheetname.strip().lower() or 'offset' in sheetname.strip().lower():
#         dfs = bg_df.index[bg_df['Unnamed: 0'].str.lower().str.contains('journ', na = False)].tolist()
#         # print(dfs)
#         cph_brut = bg_df.index[bg_df['Unnamed: 0'].str.lower().str.contains('=', na = False)].tolist()[0]
#         match = re.search(r'([0-9]+)\s*f', bg_df.at[cph_brut, 'Unnamed: 0'])
#         cph = int(match.group(1))
#         frames = []
#         for idx, val in enumerate(dfs):
#             df = pd.DataFrame()
#             st_df_end = val + 8
#             st_df = bg_df.iloc[val:st_df_end]
#             st_df = st_df.reset_index(drop=True)
#             droplist = [col for col in st_df.columns if pd.isna(st_df.at[0, col]) or 'total' in str(st_df.at[0, col]).lower()]
#             st_df.drop(droplist, axis=1, inplace=True)
#             st_df = st_df.T
#             st_df = st_df.iloc[1: , :]
#             st_df = st_df.reset_index(drop=True)
#             st_df.drop(st_df.columns[1], axis=1, inplace = True)
#             st_df.rename(columns = {
#                 st_df.columns[0]: 'date',
#                 st_df.columns[1]: 'arrets',
#                 st_df.columns[2]: 'prod_brute',
#                 st_df.columns[3]: 'shift',
#                 st_df.columns[4]: 'taux_util',
#                 st_df.columns[5]: 'capacite_prod',
#                 st_df.columns[6]: 'taux_prod',
#             }, inplace = True)
#             real = False
#             try:
#                 st_df['date'] = pd.to_datetime(st_df['date']).dt.date
#                 real = True
#             except:
#                 st_df['date'] = np.nan
#                 st_df = st_df.loc[st_df['date'].notnull()]
#             st_df = st_df.replace(['-'],0)
#             st_df.fillna(0)
#             st_df['ligne'] = sheetname
#             st_df['cph'] = cph
#             frames.append(st_df)
#         result = pd.concat(frames)
#         result = result.reset_index(drop=True)
#         cols_to_fillna = ['arrets', 'prod_brute', 'shift']
#         result[cols_to_fillna] = result[cols_to_fillna].fillna(0)
#         try:
#             if real:
#                 result.loc[result['date'] > datetime.date(2022, 4, 1), 'taux_util'] = ( (result['shift'] * 7 * 60) - result['arrets'] ) / (result['shift'] * 7 * 60)
#                 result.loc[result['date'] <= datetime.date(2022, 4, 1), 'taux_util'] = ( (result['shift'] * 8 * 60) - result['arrets'] ) / (result['shift'] * 8 * 60)
#         except:
#             result['taux_util'] = np.nan
#         try:
#             if real:
#                 result.loc[result['date'] > datetime.date(2022, 4, 1), 'capacite_prod'] = round(result['cph'] * result['shift'] * 7 * result['taux_util'])
#                 result.loc[result['date'] <= datetime.date(2022, 4, 1), 'capacite_prod'] = round(result['cph'] * result['shift'] * 8 * result['taux_util'])
#         except:
#             result['capacite_prod'] = np.nan
#         try:
#             result['taux_prod'] = result['prod_brute'] / result['capacite_prod']
#         except:
#             result['taux_prod'] = np.nan
#         bg_frames.append(result)
#     elif 'com' in sheetname.strip().lower():
#         # COMMENTAIRES, Arrets details
#         # print(bg_df)
#         dfs = bg_df.index[bg_df['Unnamed: 3'].str.lower().str.contains(r'\d{2}/\d{2}/\d{4}', na = False)].tolist()
#         frames = []
#         for idx, val in enumerate(dfs):
#             df = pd.DataFrame()
#             try:
#                 df = bg_df.loc[val: dfs[idx+1]]
#             except IndexError:
#                 df = bg_df.loc[val:]
#             dte = df['Unnamed: 3'].iloc[0]
#             # check the data type
#             if isinstance(dte, datetime.datetime):
#                 date = dte.date()
#             else:
#                 match = re.search(r'\d{2}/\d{2}/\d{4}', dte)
#                 date = datetime.datetime.strptime(match.group(), '%d/%m/%Y').date()
#             df = df.loc[:, :'Unnamed: 4']
#             df_lines = df.index[df['Unnamed: 0'].str.lower().str.contains('^vern|offs|tandem', na = False, regex = True)].tolist()
#             # print(df)
#             for sub_idx, sub_val in enumerate(df_lines):
#                 sub_df = pd.DataFrame()
#                 # print(sub_idx, sub_val)
#                 try:
#                     sub_df = df.loc[sub_val: df_lines[sub_idx+1]]
#                 except IndexError:
#                     sub_df = df.loc[sub_val:]
#                 # print(sub_df)
#                 line = sub_df['Unnamed: 0'].iloc[0]
#                 if 'ver' in line.lower():
#                     line = 'VERNISSEUSE'
#                 else:
#                     line = 'TANDEM'
#                 begin_df = (sub_df['Unnamed: 0'].str.lower().str.contains('prog', na = False)).idxmax()
#                 end_df = sub_df['Unnamed: 0'].where(sub_df['Unnamed: 0'].str.lower().str.contains('tot', na = False)).last_valid_index()
#                 sub_df = sub_df.loc[begin_df:end_df-1]
#                 sub_df = sub_df.loc[:, 'Unnamed: 1':]
#                 sub_df.rename(columns = {
#                     sub_df.columns[0]: 'label_arret',
#                     sub_df.columns[1]: 'temps arrets (mn)',
#                     sub_df.columns[2]: 'temps_prog',
#                     sub_df.columns[3]: 'taux'
#                 }, inplace = True)
#                 sub_df = sub_df[sub_df['label_arret'].notna()]
#                 sub_df['temps_prog'].fillna(method='ffill', inplace = True)
#                 sub_df['line'] = line
#                 sub_df['date'] = date
#                 sub_df['taux'] = sub_df['temps arrets (mn)'] / sub_df['temps_prog']
#                 sub_df.loc[sub_df['label_arret'].str.lower().str.contains('parat'), 'label_arret'] = 'Préparation de la ligne'
#                 sub_df.loc[sub_df['label_arret'].str.lower().str.contains('paus'), 'label_arret'] = 'Pause dejeuné'
#                 sub_df.loc[sub_df['label_arret'].str.lower().str.contains('forma'), 'label_arret'] = 'Changement de format'
#                 sub_df.loc[sub_df['label_arret'].str.lower().str.contains('lavag'), 'label_arret'] = 'Lavage'
#                 sub_df.loc[sub_df['label_arret'].str.lower().str.contains('prog'), 'label_arret'] = 'Manque de programme'
#                 sub_df.loc[sub_df['label_arret'].str.lower().str.contains('pan'), 'label_arret'] = 'Panne'
#                 sub_df.loc[sub_df['label_arret'].str.lower().str.contains('reg|rég'), 'label_arret'] = 'Réglages'
#                 sub_df.loc[sub_df['label_arret'].str.lower().str.contains('autr'), 'label_arret'] = 'Autres'
#                 sub_df.loc[sub_df['label_arret'].str.lower().str.contains('abs'), 'label_arret'] = 'Absentéisme'
#                 sub_df.fillna(0, inplace = True)
#                 frames.append(sub_df)
# # print(bg_frames)
# result = pd.concat(bg_frames)
# result = result.reset_index(drop=True)
# result = result[result['capacite_prod'].notna()]
# result_arrets = pd.concat(frames)
# result_arrets.reset_index(drop=True, inplace = True)

# result['prep_line'] = np.nan
# result['pause_eat'] = np.nan
# result['chg_form'] = np.nan
# result['lvg'] = np.nan
# result['manque_prog'] = np.nan
# result['panne'] = np.nan
# result['reglages'] = np.nan
# result['autres'] = np.nan
# result['abs'] = np.nan

# uniq_date_arret = np.sort(result_arrets['date'].unique())
# uniq_date = np.sort(result['date'].unique())
# uniq_line = result['ligne'].unique()

# full_labels = ['Préparation de la ligne',
# 'Pause dejeuné',
# 'Changement de format',
# 'Lavage',
# 'Manque de programme',
# 'Panne',
# 'Réglages',
# 'Autres',
# 'Absentéisme']

# # print(uniq_date_arret, uniq_line)

# # print(result_arrets)
# frame = []
# for dt in uniq_date_arret:
#     for line in uniq_line:
#         sub_df = result_arrets.loc[(result_arrets['date'] == dt) & (result_arrets['line'] == line)]
#         diff = list(set(full_labels) - set(sub_df['label_arret'].unique().tolist()))
#         if len(diff) > 0:
#             for label in diff:
#                 sub_df.loc[len(sub_df.index)] = [label, 0, 0, 0, line, dt]
#         # print(len(sub_df['label_arret'].unique()))
#         # frame.append(sub_df)
#         result.loc[(result['date'] == dt) & (result['ligne'] == line), 'prep_line'] = sub_df.at[sub_df[sub_df['label_arret'].str.lower().str.contains('prép')].index[0], 'temps arrets (mn)']
#         result.loc[(result['date'] == dt) & (result['ligne'] == line), 'pause_eat'] = sub_df.at[sub_df[sub_df['label_arret'].str.lower().str.contains('dej')].index[0], 'temps arrets (mn)']
#         result.loc[(result['date'] == dt) & (result['ligne'] == line), 'chg_form'] = sub_df.at[sub_df[sub_df['label_arret'].str.lower().str.contains('format')].index[0], 'temps arrets (mn)']
#         result.loc[(result['date'] == dt) & (result['ligne'] == line), 'lvg'] = sub_df.at[sub_df[sub_df['label_arret'].str.lower().str.contains('lavage')].index[0], 'temps arrets (mn)']
#         result.loc[(result['date'] == dt) & (result['ligne'] == line), 'manque_prog'] = sub_df.at[sub_df[sub_df['label_arret'].str.lower().str.contains('progr')].index[0], 'temps arrets (mn)']
#         result.loc[(result['date'] == dt) & (result['ligne'] == line), 'panne'] = sub_df.at[sub_df[sub_df['label_arret'].str.lower().str.contains('pann')].index[0], 'temps arrets (mn)']
#         result.loc[(result['date'] == dt) & (result['ligne'] == line), 'reglages'] = sub_df.at[sub_df[sub_df['label_arret'].str.lower().str.contains('régl')].index[0], 'temps arrets (mn)']
#         result.loc[(result['date'] == dt) & (result['ligne'] == line), 'autres'] = sub_df.at[sub_df[sub_df['label_arret'].str.lower().str.contains('autr')].index[0], 'temps arrets (mn)']
#         result.loc[(result['date'] == dt) & (result['ligne'] == line), 'abs'] = sub_df.at[sub_df[sub_df['label_arret'].str.lower().str.contains('abs')].index[0], 'temps arrets (mn)']

# # result_arrets = pd.DataFrame()
# # result_arrets = pd.concat(frame)
# # print(frame)
# result.reset_index(drop=True, inplace = True)

# tab_name = 'Production_Capacite_Imp'

# cursor.execute('SELECT MAX("id") FROM public."' + tab_name + '"')
# max_id = cursor.fetchone()[0]
# if not max_id:
#     max_id = 0
# result.insert(0, 'id', range(int(max_id) + 1, 1 + int(max_id) + len(result)))
# result.to_sql(tab_name, engine, if_exists='append', index=False)


# ************************** OLD CAPACITE PRODUCTION VERNISSEUSE TANDEM ****************************


