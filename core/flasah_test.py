import pandas as pd
import numpy as np
import re
import datetime
import os
from openpyxl import load_workbook

# C:\Users\zaki1\Downloads\prod_val_boites_19122021.xlsx
pd.set_option('display.max_rows', None)
pd.set_option('use_inf_as_na', True)

file_str = r'C:\Users\zaki1\Desktop\Controle de Gestion\Scripts\PRODUCTION V2\Activité journalière au 19 janvier 2022.xlsx'

def get_val_df(file_str):
    df = pd.read_excel(file_str, sheet_name='Sheet1')
    dte = df['Production valorisée (DA)'][2].date()
    df = pd.read_excel(file_str, sheet_name='Sheet1', skiprows=3).copy()
    PU_coutRev_Col = df.columns[4]
    df.columns.values[5] = 'montant_journee_coutRev'
    df.rename(columns = {
        PU_coutRev_Col: 'PU_cout_revient',
        'None.1': 'MontantCumul_coutRev',
        'Unnamed: 7': 'PU_prix_vente',
        'Unnamed: 8': 'montant_journee_prix_vente',
        'Unnamed: 9': 'MontantCumul_prix_vente',
    }, inplace = True)
    # indexNames = df[(df['Unnamed: 1'].str.contains('TOTAL', na = False))].index
    # df.drop(indexNames , inplace=True)
    subdf  = df.iloc[: , -6:]
    # print(subdf)
    return subdf

def get_prod_phy(df):
    df = df.copy()
    # df = pd.read_excel(r'C:\Users\zaki1\Downloads\prod_phy_boites_19122021.xlsx', sheet_name='Sheet1')
    dte = df['Unnamed: 6'][1]
    # print(df)
    # check the data type
    if isinstance(dte, datetime.datetime):
        date = dte.date()
    else:
        match = re.search(r'\d{2}/\d{2}/\d{4}', dte)
        date = datetime.datetime.strptime(match.group(), '%d/%m/%Y').date()

    # df = pd.read_excel(r'C:\Users\zaki1\Downloads\prod_phy_boites_19122021.xlsx', sheet_name='Sheet1', skiprows=2).copy()
    df.drop(index=df.index[:3], 
        axis=0, 
        inplace=True)
    df.drop(df.tail(1).index,inplace=True)
    # subdf = get_val_df(r'prod_val_boites_19122021.xlsx')
    # print(df.shape, subdf.shape)
    # df = pd.concat([df, subdf], axis=1)
    df['Unnamed: 0'].fillna(method='ffill', inplace = True)
    df['Unnamed: 1'].fillna(method='ffill', inplace = True)
    # df['Unnamed: 14'].fillna(method='ffill', inplace = True)
    df['date'] = date

    # OBJECTIF ET CAPACITE
    df_copy = df.copy()
    df_copy.loc[(df_copy['Unnamed: 1'].str.contains('total', na = False, flags = re.IGNORECASE)) & (~df_copy['Unnamed: 2'].isnull()), 'Unnamed: 1'] = np.nan
    df_copy.loc[df_copy['Unnamed: 1'].str.contains('total', na = False, flags = re.IGNORECASE), 'Unnamed: 0'] = np.nan
    df_copy['Unnamed: 1'] = df_copy['Unnamed: 1'].str.upper().str.strip()
    df_copy['Unnamed: 0'].fillna(method='ffill', inplace = True)
    df_copy = df_copy.loc[(df_copy['Unnamed: 1'].str.contains('total', na = False, flags = re.IGNORECASE)) & (~df_copy['Unnamed: 1'].str.contains('KDU', na = False)) & (~df_copy['Unnamed: 1'].str.contains('AZDU', na = False)) & (~df_copy['Unnamed: 1'].str.contains('SKDU', na = False)) & (~df_copy['Unnamed: 1'].str.contains('general', flags = re.IGNORECASE, na = False))]
    df_copy['Volume'] = np.nan
    df_copy.loc[(df_copy['Unnamed: 1'].str.contains('CONSE')) | (df_copy['Unnamed: 1'].str.contains('DIVER')), 'category'] = df_copy['Unnamed: 1'].str.split().str[1]
    df_copy.loc[(~df_copy['Unnamed: 1'].str.contains('CONSE')) & (~df_copy['Unnamed: 1'].str.contains('DIVER')), 'Volume'] = df_copy['Unnamed: 1'].str.partition(' ')[2]
    df_copy = df_copy.reset_index(drop=True)
    idx_cat = df_copy[df_copy['category'].notnull()].index.tolist()
    stop = 0
    end = idx_cat[-1]
    for idx in idx_cat:
        cat = df_copy.at[idx, 'category']
        for i in range(stop, idx):
            df_copy.at[i, 'category'] = cat
        stop = idx + 1
        if idx == end:
            break
    df_copy['produit'] = 'Boite'
    df_copy = df_copy.loc[df_copy['Volume'].notnull()]
    df_copy['Unnamed: 5'].fillna(0, inplace=True)
    df_copy['Unnamed: 4'].fillna(0, inplace=True)
    print(df_copy)

    indexNames = df[(df['Unnamed: 1'].str.contains('TOTAL')) & (df['Unnamed: 1'].str.contains('KDU') | df['Unnamed: 1'].str.contains('AZDU') | df['Unnamed: 1'].str.contains('SKDU') | df['Unnamed: 1'].str.contains('GENERAL'))].index
    df.drop(indexNames , inplace=True)
    df['Volume'] = np.nan
    df['category'] = np.nan
    df['produit'] = 'Boite'
    # df.loc[(df['Unnamed: 2'].notnull()) & (df['Unnamed: 1'].str.contains('TOTAL', na = False)), 'Unnamed: 1'] = np.nan
    # idx_vol = []
    # idx_cat = []
    tots = df.index[df['Unnamed: 1'].str.contains('TOTAL', na = False) | df['Unnamed: 2'].str.contains('TOTAL', na = False)].tolist()
    # [print(x, df['Unnamed: 1'].loc[x], df['Unnamed: 2'].loc[x]) for x in tots ]
    prec_vol = 0
    prec_cat = 0
    # print(tots)
    # print(df.head())
    # print(df['Unnamed: 2'])
    for idx, val in enumerate(tots):
        # print(tots[idx], df['Unnamed: 2'].loc[val])
        tot = df['Unnamed: 1'].loc[val]
        if 'total' not in tot.lower():
            # print('in if')
            tot = df['Unnamed: 2'].loc[val]
        # print(tot)
        m = re.search('.*TOTAL (.*)', tot)
        found = ''
        if m:
            found = m.group(1)
        # print(val, found)
        try:
            # print(tots[idx+1] - tots[idx])
            # print(found)
            if not any(char.isdigit() for char in found):
                # print('category')
                # print(prec_cat, val)
                for i in range(prec_cat, val):
                    try:
                        if not pd.isnull(df['Unnamed: 2'].loc[i]):
                            df['category'].loc[i] = str(found)
                    except KeyError:
                        i+=1
                # print(prec_cat)
                prec_cat = val
            else:
                # print('volume')  
                for i in range(prec_vol, val):
                    # print(i, prec_vol, val)
                    try:
                        if not pd.isnull(df['Unnamed: 2'].loc[i]) and pd.isnull(df['category'].loc[i]):
                            df['Volume'].loc[i] = str(found)
                    except KeyError:
                        i+=1
                prec_vol = val
        except IndexError:
            pass

    # rename the fields
    # print(df.columns)
    df.rename(columns = {
        'Unnamed: 0': 'Unité',
        'Unnamed: 1': 'Ligne',
        'Unnamed: 2': 'Désignation',
        'Unnamed: 3': 'Client',
        'Unnamed: 4': 'Objectif',
        'Unnamed: 5': 'Capacité jour',
        'Unnamed: 6': 'Brute_jour',
        'Unnamed: 7': 'Conforme_jour',
        'Unnamed: 8': 'Rebut_jour',
        'Unnamed: 9': 'Taux_jour',
        'Unnamed: 10': 'Brute_mois',
        'Unnamed: 11': 'Conforme_mois',
        'Unnamed: 12': 'Rebut_mois',
        'Unnamed: 13': 'Taux_real',
        'Unnamed: 14': 'Taux_rebut',
    }, inplace = True)
    df_copy.rename(columns = {
        'Unnamed: 0': 'Unité',
        'Unnamed: 1': 'Ligne',
        'Unnamed: 2': 'Désignation',
        'Unnamed: 3': 'Client',
        'Unnamed: 4': 'Objectif',
        'Unnamed: 5': 'Capacité jour',
        'Unnamed: 6': 'Brute_jour',
        'Unnamed: 7': 'Conforme_jour',
        'Unnamed: 8': 'Rebut_jour',
        'Unnamed: 9': 'Taux_jour',
        'Unnamed: 10': 'Brute_mois',
        'Unnamed: 11': 'Conforme_mois',
        'Unnamed: 12': 'Rebut_mois',
        'Unnamed: 13': 'Taux_real',
        'Unnamed: 14': 'Taux_rebut',
    }, inplace = True)

    
    # For the old ones : Rename coutrev etc prod valorisé
    # df.rename(columns = {
    #     'Unnamed: 15': 'PU_cout_revient',
    #     'Unnamed: 16': 'montant_journee_coutRev',
    #     'Unnamed: 17': 'MontantCumul_coutRev',
    #     'Unnamed: 18': 'PU_prix_vente',
    #     'Unnamed: 19': 'montant_journee_prix_vente',
    #     'Unnamed: 20': 'MontantCumul_prix_vente'
    # }, inplace = True)


    df = df[df['Désignation'].notna()]
    df.insert(15, 'PU_cout_revient', np.nan)
    df.insert(16, 'montant_journee_coutRev', np.nan)
    df.insert(17, 'MontantCumul_coutRev', np.nan)   
    df.insert(18, 'PU_prix_vente', np.nan)
    df.insert(19, 'montant_journee_prix_vente', np.nan)
    df.insert(20, 'MontantCumul_prix_vente', np.nan)
    # print(df.dtypes)

    # delete rows where total exist in client
    indexNames = df[(df['Client'].str.contains('TOTAL', na = False))].index
    df.drop(indexNames , inplace=True)
    indexNames = df[(df['Ligne'].str.contains('TOTAL', na = False))].index
    df.drop(indexNames , inplace=True)

    # w['female'] = w['female'].map({'female': 1, 'male': 0})
    
    # For the old ones tenik
    # print(df.iloc[:, : 8])
    # wrap_idx = df.index[df['Ligne'].str.lower().str.contains('total', na = False)].tolist()[0]
    # df = df.loc[:wrap_idx-3]
    # df['Taux_real'] = np.array(df['Taux_real'], dtype=float)
    # df['Taux_rebut'] = np.array(df['Taux_rebut'], dtype=float)
    # df['Taux_jour'] = np.array(df['Taux_jour'], dtype=float)
    
    # print(df[['Rebut_mois', 'Brute_mois']].head())
    # try:
    df.loc[df['Rebut_mois'].astype(str).str.isspace(), 'Rebut_mois'] = np.nan
    df.loc[df['Brute_mois'].astype(str).str.isspace(), 'Brute_mois'] = np.nan
    df.loc[df['Objectif'].astype(str).str.isspace(), 'Objectif'] = np.nan
    df.loc[df['Brute_jour'].astype(str).str.isspace(), 'Brute_jour'] = np.nan
    df.loc[df['Capacité jour'].astype(str).str.isspace(), 'Capacité jour'] = np.nan
    df['Taux_rebut'] = df['Rebut_mois'] / df['Brute_mois'].replace({ 0 : np.nan })
    # except ZeroDivisionError:
    #     df['Taux_rebut'] = 0
    # try:
    df['Taux_real'] = df['Brute_mois'] / df['Objectif'].replace({ 0 : np.nan })
    # except ZeroDivisionError: 
    #     df['Taux_real'] = 0
    # # df['Capacité jour'] = 0
    # try:
    df['Taux_jour'] = df['Brute_jour'] / df['Capacité jour'].replace({ 0 : np.nan })
    # except ZeroDivisionError:
    #     df['Taux_jour'] = 0
    # df.loc[np.isinf(df['Taux_real']), 'Taux_real'] = 0
    # df.loc[np.isinf(df['Taux_rebut']), 'Taux_rebut'] = 0
    # df.loc[np.isinf(df['Taux_jour']), 'Taux_jour'] = 0
    if date.month < 10:
        month = '0' + str(date.month)
    else:
        month = str(date.month)
    df = df.replace(['-'],0)
    df.loc[df['Unité'] == 'AZDU', 'category'] = 'CONSERVE'
    df.loc[df['Unité'] == 'SKDU', 'category'] = 'DIVERS'
    for col in df.columns:
        if col != 'Client' and col != 'Volume' and col != 'Désignation':
            df[col].fillna(0, inplace=True)
    indexNames = df[df['Ligne'].str.lower().str.contains('total', na = False) | df['Désignation'].str.lower().str.contains('total', na = False)].index
    df.drop(indexNames , inplace=True)
    # file_name = 'Flash_Journ_PROD_PHY_' + month + '_' + str(date.year) + '.xlsx'
    # if not os.path.isfile(r'C:\Users\zaki1\Desktop\Controle de Gestion\Scripts\Flash_Journ_PROD_PHY_' + month + '_' + str(date.year) + '.xlsx'):
    #     # print('New file !')
    #     df.to_excel(file_name, index = False)
    # else:
    #     # print('file exists !')
    #     book = load_workbook(file_name)
    #     writer = pd.ExcelWriter(file_name, engine='openpyxl', mode='a') # pylint: disable=abstract-class-instantiated
    #     writer.book = book
    #     writer.sheets = {ws.title: ws for ws in book.worksheets}
    #     for sheetname in writer.sheets:
    #         # print('for sheetname')
    #         # print(sheetname, writer.sheets[sheetname].max_row)
    #         df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
    #     writer.save()
    # print(df)
    # print(date)
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    print(df.shape)
    # print(df.head())
    # print(df['Taux_real'])
    return df, df_copy

if __name__ == '__main__':
    bg_df = pd.read_excel(file_str, sheet_name='01 Prod Physique Boites', skiprows=7, header=1)
    # Rename the first column so we could address it
    bg_df.columns.values[0] = 'Unnamed: 0'
    bg_df.columns.values[1] = 'Unnamed: 1'
    bg_df.columns.values[2] = 'Unnamed: 2'
    bg_df.columns.values[3] = 'Unnamed: 3'
    bg_df.columns.values[6] = 'Unnamed: 6'

    # For the old ones (september w etle3)
    # bg_df = bg_df.iloc[: , :21]
    # bg_df.loc[bg_df['Unnamed: 0'].str.lower().str.contains(r'^\s*unité\s*kdu\s*$', na = False, regex = True), 'Unnamed: 0'] = 'KDU'
    # bg_df.loc[bg_df['Unnamed: 0'].str.lower().str.contains(r'^\s*unité\s*azdu\s*$', na = False, regex = True), 'Unnamed: 0'] = 'AZDU'
    # bg_df.loc[bg_df['Unnamed: 0'].str.lower().str.contains(r'^\s*skdu\s*$', na = False, regex = True), 'Unnamed: 0'] = 'SKDU'

    dfs = bg_df.index[bg_df['Unnamed: 0'].str.contains('Unité', na = False)].tolist()
    for idx, val in enumerate(dfs):
        df = pd.DataFrame()
        # print(val, dfs[idx+1])
        try:
            df = bg_df.loc[val: dfs[idx+1]]
        except IndexError:
            df = bg_df.loc[val:]
        # df = df.iloc[3: , :]
        df = df.reset_index(drop=True)
        df.drop(df.tail(1).index,inplace=True)
        # print(df.iloc[:, : 8])
        # print('Done')
        # if last line, then remove last meaningless lines
        # if val == dfs[-1]:
        df = get_prod_phy(df)
        # break
        # print(df)

# label end

# Comment the fonction section in order to add the ID column

# df.insert(0, 'ID', range(int(max_id) + 1, 1 + int(max_id) + len(df)))
