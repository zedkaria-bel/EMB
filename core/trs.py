import pandas as pd
import numpy as np
import re
import datetime
import os
from openpyxl import load_workbook

pd.set_option('display.max_rows', None)

def get_trs_df(df):
    df = df.copy()
    dte = df['Unnamed: 5'][0]
    if isinstance(dte, datetime.datetime):
        date = dte.date()
    else:
        match = re.search(r'\d{2}/\d{2}/\d{4}', dte)
        date = datetime.datetime.strptime(match.group(), '%d/%m/%Y').date()
    # print(date)
    df.drop(index=df.index[:2], 
        axis=0, 
        inplace=True)
    df.loc[df['Unnamed: 0'].astype(str).str.contains('5', na = False), 'Unnamed: 0'] = 'KDU'
    # print(df['Unnamed: 0'])
    df['Unnamed: 0'].fillna(method='ffill', inplace = True)
    df['date'] = date
    df.rename(columns = {
        'Unnamed: 0': 'Unité',
        'Unnamed: 1': 'Ligne',
        'Unnamed: 2': 'Ecarts_Cadences',
        'Unnamed: 3': 'Arret_Plan',
        'Unnamed: 4': 'Arret_non_Plan',
        'Unnamed: 5': 'Capacite_Theo',
        'Unnamed: 6': 'Qte_Prod',
        'Unnamed: 7': 'Qte_Conf',
        'Unnamed: 8': 'Qte_Rebut',
        'Unnamed: 9': 'Temps_Ouv',
        'Unnamed: 10': 'Temps_Fct',
        'Unnamed: 11': 'Temps_Req',
        'Unnamed: 12': 'Taux_Dispo',
        'Unnamed: 13': 'Temps_Net',
        'Unnamed: 14': 'Temps_fct2',
        'Unnamed: 15': 'Taux_Perf',
        'Unnamed: 16': 'Temps_Util',
        'Unnamed: 17': 'Temps_net2',
        'Unnamed: 18': 'Taux_Qualit',
        'Unnamed: 19': 'TRS',
    }, inplace = True)
    indexNames = df[df['Unité'].str.contains('Unité', na = False)].index
    df.drop(indexNames , inplace=True)
    df = df[df['Ligne'].notna()]
    df = df.replace(['-'],0)
    df.fillna(0, inplace=True)
    # print(df.columns)
    df.drop(['Temps_fct2', 'Temps_net2'], axis=1, inplace = True)
    # file_name = 'TRS_' + str(date.month) + '_' + str(date.year) + '.xlsx'
    # if not os.path.isfile(r'C:\Users\zaki1\Desktop\Controle de Gestion\Scripts\TRS_' + str(date.month) + '_' + str(date.year) + '.xlsx'):
    #     # print('New file !')
    #     df.to_excel(file_name, index = False)
    # else:
    #     # print('file exists !')
    #     book = load_workbook(file_name)
    #     writer = pd.ExcelWriter(file_name, engine='openpyxl', mode='a') # pylint: disable=abstract-class-instantiated
    #     writer.book = book
    #     writer.sheets = {ws.title: ws for ws in book.worksheets}
    #     for sheetname in writer.sheets:
    #         # print('for sheetname')
    #         # print(sheetname, writer.sheets[sheetname].max_row)
    #         df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
    #     writer.save()
    print(df.shape)
    # print(df.iloc[:, :6])
    return df

file_str = r'C:\Users\zaki1\Desktop\Controle de Gestion\Scripts\PRODUCTION V2\Activité journalière au 19 janvier 2022.xlsx'
if __name__ == '__main__':
    bg_df = pd.read_excel(file_str, sheet_name='06 TRS', skiprows=8, header=1)
    bg_df.columns.values[0] = 'Unnamed: 0'
    bg_df.columns.values[1] = 'Unnamed: 1'
    dfs = bg_df.index[bg_df['Unnamed: 0'].str.contains('Unité', na = False)].tolist()
    for idx, val in enumerate(dfs):
        df = pd.DataFrame()
        try:
            df = bg_df.loc[val: dfs[idx+1]]
        except IndexError:
            df = bg_df.loc[val:]
        df = df.reset_index(drop=True)
        df = get_trs_df(df)
# print(bg_df)




# Comment the fonction section in order to add the ID column

# df.insert(0, 'ID', range(int(max_id) + 1, 1 + int(max_id) + len(df)))

