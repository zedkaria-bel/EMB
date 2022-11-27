import pandas as pd
import numpy as np
import re
import datetime
import os
from openpyxl import load_workbook

pd.set_option('display.max_rows', None)

def get_vente_df(df):
    df = df.copy()
    dte = df['Unnamed: 4'][0]
    if isinstance(dte, datetime.datetime):
        date = dte.date()
    else:
        match = re.search(r'\d{2}/\d{2}/\d{4}', dte)
        date = datetime.datetime.strptime(match.group(), '%d/%m/%Y').date()
    df.drop(index=df.index[:4], 
        axis=0, 
        inplace=True)
    df['Unnamed: 0'].fillna(method='ffill', inplace = True)
    df['date'] = date
    df['category'] = np.nan
    # df.drop(['Unnamed: 9', 'Unnamed: 10'], axis=1, inplace = True)
    df.rename(columns = {
        'Unnamed: 0': 'Unité',
        'Unnamed: 1': 'Ligne',
        'Unnamed: 2': 'Désignation',
        'Unnamed: 3': 'Client',
        'Unnamed: 4': 'Qte_Journ',
        'Unnamed: 5': 'Qte_Cumul',
        'Unnamed: 6': 'PU',
        'Unnamed: 7': 'Montant_journee',
        'Unnamed: 8': 'Montant_cumul',
    }, inplace = True)
    indexNames = df[(~df['Ligne'].isnull()) & ((~df['Ligne'].str.contains('CONSERVE', na = False)) & (~df['Ligne'].str.contains('DIVERS', na = False)) & (~df['Ligne'].str.contains('AUTRE', na = False)))].index
    df.drop(indexNames , inplace=True)
    indexNames = df[df['Ligne'].isnull() & df['Désignation'].isnull() & df['Client'].isnull()].index
    df.drop(indexNames , inplace=True)
    # print(df)
    tots = df.index[df['Ligne'].str.contains('TOTAL|AUTRES', na = False, regex = True)].tolist()
    # print(tots)
    prec_cat = 0
    for idx, val in enumerate(tots):
        tot = df['Ligne'].loc[val]
        # print(tot)
        m = re.search(r'(.*TOTAL (I|II)\s+(\w+)\s*)|(AUTRES)', tot)
        # print(m.groups())
        found = ''
        if m:
            if m.group(4) is None:
                found = m.group(3)
            else:
                found = m.group(4)
        # print(found)
        try:
            # print(tots[idx+1] - tots[idx])
            # print(found)
            for i in range(prec_cat, val):
                try:
                    if not pd.isnull(df['Montant_journee'].loc[i]):
                        # print(tot, df['Unnamed: 2'].loc[i])
                        # print(found)
                        df['category'].loc[i] = str(found)
                except KeyError:
                    i+=1
            # print(prec_cat)  
            prec_cat = val
        except IndexError:
            pass
    df = df.replace(['-'],0)
    indexNames = df[df['Ligne'].str.contains('TOTAL', na = False) | df['Ligne'].str.contains('AUTRES', na = False)].index
    df.drop(indexNames , inplace=True)
    df.drop(['Ligne'], axis=1, inplace = True)
    df = df[df['category'].notna()]
    for col in df.columns:
        if col != 'Client' and col != 'Désignation':
            df[col].fillna(0, inplace=True)
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

    # print('XXXX')
    # print(df['Montant_journee'])
    # print(df['Qte_Journ'])
    # print(df['Qte_Cumul'])
    # print(df['PU'])
    df.loc[(df['Montant_journee'] != 0) & ((df['Qte_Journ'] == 0) | (df['Qte_Cumul'].isnull())), 'Qte_Journ'] = df['Montant_journee'] / df['PU']
    print('PPPP')
    df.loc[(df['Qte_Cumul'] == 0) & (df['Qte_Journ'] != 0), 'Qte_Cumul'] = df['Qte_Journ']
    indexNames = df[df['Qte_Cumul'] == 0].index
    df.drop(indexNames , inplace=True)
    df.loc[df['category'].str.contains('diver', flags=re.IGNORECASE, na = False), 'category'] = 'DIVERSE'
    df['Unité'] = df['Unité'].str.upper()
    print(df)
    # file_name = 'Ventes_' + str(date.month) + '_' + str(date.year) + '.xlsx'
    # if not os.path.isfile(r'C:\Users\zaki1\Desktop\Controle de Gestion\Scripts\Ventes_' + str(date.month) + '_' + str(date.year) + '.xlsx'):
    #     # print('New file !')
    #     df.to_excel(file_name, index = False)
    # else:
    #     # print('file exists !')
    #     book = load_workbook(file_name)
    #     writer = pd.ExcelWriter(file_name, engine='openpyxl', mode='a') # pylint: disable=abstract-class-instantiated
    #     writer.book = book
    #     writer.sheets = {ws.title: ws for ws in book.worksheets}
    #     for sheetname in writer.sheets:
    #         # print('for sheetname')
    #         # print(sheetname, writer.sheets[sheetname].max_row)
    #         df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
    #     writer.save()
    print(df.shape)
    return df

file_str = r'C:\Users\zaki1\Desktop\Controle de Gestion\Scripts\PRODUCTION V2\Activité journalière au 19 janvier 2022.xlsx'
if __name__ =='__main__':
    bg_df = pd.read_excel(file_str, sheet_name='04 Ventes', skiprows=8, header=1)
    bg_df.columns.values[0] = 'Unnamed: 0'
    bg_df.columns.values[1] = 'Unnamed: 1'
    bg_df.columns.values[2] = 'Unnamed: 2'
    bg_df.columns.values[3] = 'Unnamed: 3'
    dfs = bg_df.index[bg_df['Unnamed: 0'].str.contains('Unité', na = False)].tolist()
    for idx, val in enumerate(dfs):
        df = pd.DataFrame()
        try:
            df = bg_df.loc[val: dfs[idx+1]]
        except IndexError:
            df = bg_df.loc[val:]
        df = df.reset_index(drop=True)
        df = get_vente_df(df)

# print(bg_df.head())


# Comment the fonction section in order to add the ID column

# df.insert(0, 'ID', range(int(max_id) + 1, 1 + int(max_id) + len(df)))