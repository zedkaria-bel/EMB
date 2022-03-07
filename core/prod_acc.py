import pandas as pd
import numpy as np
import re
import datetime
import os
from openpyxl import load_workbook

def get_acc_df(df):
    # df = pd.read_excel(file_str, sheet_name='Feuil1')
    dte = df['Unnamed: 5'][1]
    # print(dte)
    if isinstance(dte, datetime.datetime):
        date = dte.date()
    else:
        match = re.search(r'\d{2}/\d{2}/\d{4}', dte)
        date = datetime.datetime.strptime(match.group(), '%d/%m/%Y').date()
    # df = pd.read_excel(file_str, sheet_name='03 Prod Accessoires', skiprows=2).copy()
    df.drop(index=df.index[:3], 
        axis=0, 
        inplace=True)
    # print(df.iloc[:, : 6])
    indexNames = df[df['Unnamed: 1'].isnull() & df['Unnamed: 2'].isnull()].index
    df.drop(indexNames , inplace=True)
    df['Unnamed: 0'].fillna(method='ffill', inplace = True)
    df['Unnamed: 1'].fillna(method='ffill', inplace = True)
    df['date'] = date
    # print(df)

    # OBJECTIF ET CAPACITE
    df_copy = df.copy()
    df_copy.loc[(df_copy['Unnamed: 1'].str.contains('total', na = False, flags = re.IGNORECASE)) & (~df_copy['Unnamed: 2'].isnull()), 'Unnamed: 1'] = np.nan
    df_copy.loc[df_copy['Unnamed: 1'].str.contains('total', na = False, flags = re.IGNORECASE), 'Unnamed: 0'] = np.nan
    df_copy.loc[(~df_copy['Unnamed: 1'].str.contains('total', na = False, flags = re.IGNORECASE) & (~df_copy['Unnamed: 1'].isnull())), 'Unnamed: 0'] = 'SKDU'
    df_copy['Unnamed: 0'].fillna(method='ffill', inplace = True)
    df_copy = df_copy.loc[(df_copy['Unnamed: 1'].str.contains('total', na = False, flags = re.IGNORECASE)) & (~df_copy['Unnamed: 1'].str.contains('KDU', na = False)) & (~df_copy['Unnamed: 1'].str.contains('AZDU', na = False)) & (~df_copy['Unnamed: 1'].str.contains('SKDU', na = False)) & (~df_copy['Unnamed: 1'].str.contains('general', flags = re.IGNORECASE, na = False))]
    df_copy['Volume'] = np.nan
    df_copy['category'] = df_copy['Unnamed: 1'].str.split().str[1]
    df_copy['produit'] = 'Accessoire'

    indexNames = df[(df['Unnamed: 1'].str.contains('TOTAL')) & (df['Unnamed: 1'].str.contains('KDU') | df['Unnamed: 1'].str.contains('AZDU') | df['Unnamed: 1'].str.contains('SKDU') | df['Unnamed: 1'].str.contains('GENERAL')) & df['Unnamed: 2'].isnull()].index
    df.drop(indexNames , inplace=True)
    df['Volume'] = np.nan
    df['category'] = np.nan
    df['produit'] = 'Accessoire'
    df.loc[(df['Unnamed: 2'].notnull()) & (df['Unnamed: 1'].str.contains('TOTAL', na = False)), 'Unnamed: 1'] = np.nan
    # print(df.head())
    # idx_vol = []
    # idx_cat = []
    tots = df.index[df['Unnamed: 1'].str.contains('TOTAL', na = False)].tolist()
    prec_vol = 0
    prec_cat = 0
    # print(tots)
    for idx, val in enumerate(tots):
        # print(tots[idx], df['Unnamed: 2'].loc[val])
        tot = df['Unnamed: 1'].loc[val]
        # print(tot)
        m = re.search('.*TOTAL (.*)', tot)
        found = ''
        if m:
            found = m.group(1)
        # print(val, found)
        try:
            # print(tots[idx+1] - tots[idx])
            # print(found)
            if not any(char.isdigit() for char in found):
                for i in range(prec_cat, val):
                    try:
                        if not pd.isnull(df['Unnamed: 2'].loc[i]):
                            # print(tot, df['Unnamed: 2'].loc[i])
                            df['category'].loc[i] = str(found)
                    except KeyError:
                        i+=1
                # print(prec_cat)  
                prec_cat = val
            else:
                # print('volume')  
                for i in range(prec_vol, val):
                    # print(i, prec_vol, val)
                    try:
                        if not pd.isnull(df['Unnamed: 2'].loc[i]):
                            df['Volume'].loc[i] = str(found)
                    except KeyError:
                        i+=1
                prec_vol = val
        except IndexError:
            pass
    # print(df.columns)
    df.rename(columns = {
        'Unnamed: 0': 'Unité',
        'Unnamed: 1': 'Ligne',
        'Unnamed: 2': 'Désignation',
        'Unnamed: 3': 'Objectif',
        'Unnamed: 4': 'Capacité jour',
        'Unnamed: 5': 'Brute_jour',
        'Unnamed: 6': 'Conforme_jour',
        'Unnamed: 7': 'Rebut_jour',
        'Unnamed: 8': 'Taux_jour',
        'Unnamed: 9': 'Brute_mois',
        'Unnamed: 10': 'Conforme_mois',
        'Unnamed: 11': 'Rebut_mois',
        'Unnamed: 12': 'Taux_real',
        'Unnamed: 13': 'Taux_rebut',
        'Unnamed: 14': 'PU_cout_revient',
        'Unnamed: 15': 'montant_journee_coutRev',
        'Unnamed: 16': 'MontantCumul_coutRev',
    }, inplace = True)
    df_copy.rename(columns = {
        'Unnamed: 0': 'Unité',
        'Unnamed: 1': 'Ligne',
        'Unnamed: 2': 'Désignation',
        'Unnamed: 3': 'Objectif',
        'Unnamed: 4': 'Capacité jour',
        'Unnamed: 5': 'Brute_jour',
        'Unnamed: 6': 'Conforme_jour',
        'Unnamed: 7': 'Rebut_jour',
        'Unnamed: 8': 'Taux_jour',
        'Unnamed: 9': 'Brute_mois',
        'Unnamed: 10': 'Conforme_mois',
        'Unnamed: 11': 'Rebut_mois',
        'Unnamed: 12': 'Taux_real',
        'Unnamed: 13': 'Taux_rebut',
        'Unnamed: 14': 'PU_cout_revient',
        'Unnamed: 15': 'montant_journee_coutRev',
        'Unnamed: 16': 'MontantCumul_coutRev',
    }, inplace = True)
    df = df.replace(['-'],0)
    df_copy = df_copy.replace(['-'],0)
    for col in df.columns:
        if col != 'Ligne' and col != 'Volume' and col != 'Désignation':
            df[col].fillna(0, inplace=True)
    for col in df_copy.columns:
        if col != 'Ligne' and col != 'Volume' and col != 'Désignation':
            df_copy[col].fillna(0, inplace=True)
    df.loc[df['Ligne'].str.contains('TOTAL', na = False), 'Ligne'] = np.nan
    df = df[df['Désignation'].notna()]
    # if date.day < 10:
    #     day = '0' + str(date.day)

    # df['Client'] = np.nan
    # df['PU_prix_vente'] = np.nan
    # df['montant_journee_prix_vente'] = np.nan
    # df['MontantCumul_prix_vente'] = np.nan
    # print(file_name)
    # print(df.columns)
    # print(date)
    # print(df.iloc[:, : 6])
    df.insert(3, 'Client', np.nan)
    df.insert(18, 'PU_prix_vente', 0)
    df.insert(19, 'montant_journee_prix_vente', 0)
    df.insert(20, 'MontantCumul_prix_vente', 0)

    df_copy.insert(3, 'Client', np.nan)
    df_copy.insert(18, 'PU_prix_vente', 0)
    df_copy.insert(19, 'montant_journee_prix_vente', 0)
    df_copy.insert(20, 'MontantCumul_prix_vente', 0)

    # For the new files only
    df.loc[df['Ligne'].notna(), 'Unité'] = 'SKDU'


    if date.month < 10:
        month = '0' + str(date.month)
    else:
        month = str(date.month)
    # file_name = 'Flash_Journ_ACC_' + month + '_' + str(date.year) + '.xlsx'
    # if not os.path.isfile(r'C:\Users\zaki1\Desktop\Controle de Gestion\Scripts\Flash_Journ_ACC_' + month + '_' + str(date.year) + '.xlsx'):
    #     # print('New file !')
    #     df.to_excel(file_name, index = False)
    # else:
    #     # print('file exists !')
    #     book = load_workbook(file_name)
    #     writer = pd.ExcelWriter(file_name, engine='openpyxl', mode='a') # pylint: disable=abstract-class-instantiated
    #     writer.book = book
    #     writer.sheets = {ws.title: ws for ws in book.worksheets}
    #     for sheetname in writer.sheets:
    #         # print('for sheetname')
    #         # print(sheetname, writer.sheets[sheetname].max_row)
    #         df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
    #     writer.save()
    # print(df)
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    df_copy = df_copy.loc[:, ~df_copy.columns.str.contains('^Unnamed')]
    print(df.shape)
    return df, df_copy

# df = get_acc_df(r'C:\Users\zaki1\Downloads\prod_acc_19122021.xlsx')

# Read the entire excel file

file_str = r'C:\Users\zaki1\Desktop\Controle de Gestion\Scripts\PRODUCTION V2\Activité journalière au 19 janvier 2022.xlsx'

# date = datetime.date(2021, 12, 19)
# print(str(date.day) + str(date.month) + str(date.year))

if __name__ == '__main__':
    bg_df = pd.read_excel(file_str, sheet_name='03 Prod Accessoires', skiprows=8, header=1)
    # Rename the first column so we could address it
    bg_df.columns.values[0] = 'Unnamed: 0'
    bg_df.columns.values[1] = 'Unnamed: 1'
    bg_df.columns.values[2] = 'Unnamed: 2'
    # print(bg_df.iloc[:, : 8])

    pd.set_option('display.max_rows', None)
    # # Delimiter les dataframes
    dfs = bg_df.index[bg_df['Unnamed: 0'].str.contains('Unité', na = False)].tolist()
    # print(dfs)
    # dfs_list = list(enumerate(dfs))
    # print(dfs_list)
    for idx, val in enumerate(dfs):
        # print(idx, val)
        df = pd.DataFrame()
        # print(val, dfs[idx+1])
        try:
            df = bg_df.loc[val: dfs[idx+1]]
        except IndexError:
            df = bg_df.loc[val:]
        # df = df.iloc[3: , :]
        df = df.reset_index(drop=True)

        # Part for the old ones
        # To comment if trying new model : calcul taux de rebut
        # Add new column at position 11 starting from 0 in the excel file (from unité)
        # df.insert(11, 'Taux de Rebut', )

        df.drop(df.tail(1).index,inplace=True)
        # print(df.iloc[:, : 8])
        # # if last line, then remove last meaningless lines
        # if val == dfs[-1]:
        
        df = get_acc_df(df)
        # print(df)
        # break
    # print(dfs)

# print(df.head(50))

# Comment the fonction section in order to add the ID column

# df.insert(0, 'ID', range(int(max_id) + 1, 1 + int(max_id) + len(df)))